# Yuzhou Huang
#
# IMPORTANT: I REWROTE course_dictionary.py before I realized that it would affect grading. Unfortunately, i realized my
# mistake too late in the process and I would have to rewrite everything in yuzhouhuang_scheduler to compensate.
#
# The code uses a naive depth-first regression search to find a schedule that satisfies certain goal requirements,
# given an arbitrary set of preconditions.
#
# The code stores a dictionary of all the courses and their information, and uses said dictionary to build a list of
# possible operators. The code does not store a graph in data. Rather, to find the children of a given node, it applies
# any operator with an effect that creates some or all of the current node state.
#
# I tested
#   gs: CS1101 in:
#   gs: CS2201 in: CS1101
#   gs: CS3251 in:
#   gs: CSmajor
#   gs: 'CS', '2231' 'CS', '3251' 'CS', 'statsprobability' in: 'MATH' '2810' 'MATH' '2820' 'MATH' '3640'

import re
from collections import namedtuple
from openpyxl import load_workbook

Course = namedtuple('Course', 'program, designation')
CourseInfo = namedtuple('CourseInfo', 'credits, terms, prereqs')
Operator = namedtuple('Operator', 'pre, eff, ScheduledTerm, credits')
semesters = ['Fall', 'Spring', 'Summer']
Term = namedtuple("Term", "year, semester")
years = ['Frosh', 'Soph', 'Junior', 'Senior']


# Top level, big bad guy
# Eats garbage and shoots out a beautiful schedule
def course_scheduler(course_descriptions, goal_conditions, initial_state):
    basic_plan = dict()
    operators = create_course_operators(course_descriptions)

    blah_blah = get_operators(operators, goal_conditions)
    for op in blah_blah:
        if dfs(operators, goal_conditions, initial_state, op, basic_plan):
            return polish_plan(basic_plan, course_descriptions)

    return ()


# Applies an operator to a statte
def apply_regression_operator(state, operator):
    if operator and operator.eff in state:
        new_state.remove(operator.eff)
        new_state.extend(operator.pre)
    return new_state


# Creates a dictionary -- modified to change some tuples into lists
def create_course_dict():
    """
    Creates a dictionary containing course info.
    Keys: namedtuple of the form ('program, designation')
    Values: namedtuple of the form('name, prereqs, credits')
            prereqs is a list of prereqs where each prereq has the same form as the keys
    """
    wb = load_workbook('newcatalog.xlsx', data_only=True)
    catalog = wb.get_sheet_by_name('catalog')
    course_dict = {}
    for row in range(1, catalog.max_row + 1):
        key = Course(get_val(catalog, 'A', row), get_val(catalog, 'B', row))
        prereqs = list(list(get_split_course(prereq) for prereq in prereqs.split())
                       for prereqs in none_split(get_val(catalog, 'E', row)))
        val = CourseInfo(int(get_val(catalog, 'C', row)), list(get_val(catalog, 'D', row).split()), prereqs)
        course_dict[key] = val
    return course_dict


# Creates a list of operators
def create_course_operators(course_descriptions):
    course_operators = []

    for course in course_descriptions:
        course_info = course_descriptions[course]

        # New operator for each year and term
        for term in course_info.terms:
            for year in years:
                if course_info.prereqs:
                    # New operator for each set of prereqs
                    for prereq in course_info.prereqs:
                        new_op = Operator(prereq, course, Term(year, term), course_info.credits)
                        course_operators.append(new_op)
                else:
                    new_op = Operator([], course, Term(year, term), course_info.credits)
                    course_operators.append(new_op)

    return course_operators


# The meat of the operation. DFS for the correct schedule
def dfs(operators, prev_state, initial_state, operator, plan):
    new_state = apply_regression_operator(prev_state, operator)

    # Add course to tentative plan
    new_course = operator.eff
    new_course_info = CourseInfo(operator.credits, operator.ScheduledTerm, operator.pre)
    plan[new_course] = new_course_info

    # If the state has the same content as the goal, we've done it
    # For reasons of bad coding, we need to add the initial state to the new state

    if set(new_state).union(initial_state) == set(initial_state):
        return True
    else:
        new_operators = get_operators(operators, new_state)

        for new_op in new_operators:

            # If the operator creates a new class, we must worry about timing
            new_credits = new_op.credits
            if new_credits != 0:

                # If all the prereqs take place before the effect
                if prereq_is_valid(new_op, plan):
                    credit_count = 0
                    for key in plan:
                        if plan[key].terms == new_op.ScheduledTerm:
                            credit_count += plan[key].credits
                    if (new_op.ScheduledTerm.semester == 'Summer' and credit_count + new_credits <= 6) or (
                                    new_op.ScheduledTerm.semester != 'Summer' and credit_count + new_credits <= 21):
                        # if we find the goal, stop
                        if dfs(operators, new_state, initial_state, new_op, plan):
                            return True

            # Else, if the operator creates a none-class effect (e.g. a major), no need to worry about timin
            else:
                if dfs(operators, new_state, initial_state, new_op, plan):
                    return True

        # If none of the children return true, we must backtrack
        if new_course in plan:
            del plan[new_course]
        return False


# Gets the operators that create some or all of a state
def get_operators(operators, state):
    # Get each operator that could produce the state\
    output = []
    for op in operators:
        if op.eff in state:
            output.append(op)
    return output


# i dunno
def get_split_course(course):
    """
    Parses a course from programdesignation into the ('program, designation') form.
    e.g. 'CS1101' -> ('CS', '1101')
    """
    temp = tuple(split_course for course_part in re.findall('((?:[A-Z]+-)?[A-Z]+)(.+)', course)
                 for split_course in course_part)
    return Course(temp[0], temp[1])


def get_val(catalog, col, row):
    """Returns the value of a cell."""
    return catalog[col + str(row)].value


def none_split(val):
    """Handles calling split on a None value by returning the empty list."""
    return val.split(', ') if val else ()


# Fills in semesters with too few credits
def polish_plan(basic_plan, course_descriptions):
    # Stores the credit per semester
    credit_counter = dict();
    for y in years:
        for s in semesters:
            credit_counter[Term(y, s)] = 0
    for course in basic_plan:
        course_info = basic_plan[course]
        credit_counter[course_info.terms] = course_info.credits

    # if there are any subcredit semester, fill them in
    for term in credit_counter:
        # Don't need to fill in empy semesters or summer
        while credit_counter[term] != 0 and term.semester != 'Summer' and credit_counter[term] < 12:
            # get a no pre-req class in the same semester
            new_course = next(c for c in course_descriptions
                              # The course has no prereqs
                              if course_descriptions[c].prereqs == [] and
                              # The course has the same semester
                              term.semester in course_descriptions[c].terms and
                              c not in basic_plan)
            new_course_info = CourseInfo(course_descriptions[new_course].credits, term, [])
            basic_plan[new_course] = new_course_info
            credit_counter[term] += course_descriptions[new_course].credits

    final_plan = basic_plan

    return final_plan


# determines whether an operator is valid
def prereq_is_valid(operator, plan):
    # If a prereq appears in the same semester or after a course, the operation is invalid
    for p in operator.pre:
        if p in plan:
            p_term = plan[p].terms
            if term_less_than_or_equal(operator.ScheduledTerm, p_term):
                return False

    return True


def print_dict(dict):
    """Simply prints a dictionary's key and values line by line."""
    for key in dict:
        print(key, dict[key])


# Determines which term comes first
def term_less_than_or_equal(term_1, term_2):
    return years.index(term_1.year) < years.index(term_2.year) or (
        years.index(term_1.year) == years.index(term_2.year) and semesters.index(term_1.semester) <= semesters.index(
            term_2.semester))
